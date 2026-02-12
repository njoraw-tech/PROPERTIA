from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import csv
import json

# Create your views here.


from .models import Unit
from tenants.models import Tenant
from properties.models import Property
from django import forms

class UnitForm(forms.ModelForm):
    class Meta:
        model = Unit
        fields = ['property', 'name', 'rent_amount', 'description']
        widgets = {
            'property': forms.Select(attrs={'class': 'form-select'}),
            'name': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'e.g. House 712'}),
            'rent_amount': forms.NumberInput(attrs={'class': 'form-control'}),
            'description': forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
        }

def units_list(request):
    # 1. Handle POST (Form Submission)
    if request.method == "POST":
        form = UnitForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('units:units_list')
        # If form is NOT valid, we don't reset it; we let it pass 
        # to the template so the user can see the error messages.
    else:
        # 2. Handle GET (Initial Page Load)
        form = UnitForm()

    # 3. Filtering Logic (Common to both or just GET)
    units = Unit.objects.all()
    property_filter = request.GET.get('property')
    status_filter = request.GET.get('status')
    
    if property_filter:
        units = units.filter(property_id=property_filter)
    if status_filter:
        units = units.filter(status=status_filter)
            
    # 4. Return the Response
    return render(request, 'units/units_view.html', {
        'units': units,
        'form': form, # This is now guaranteed to exist!
        'properties': Property.objects.all(),
        'tenants': Tenant.objects.filter(unit__isnull=True),
    })


def assign_tenant(request, pk):
    """Assign a tenant to a unit"""
    unit = get_object_or_404(Unit, pk=pk)
    
    if request.method == 'POST':
        tenant_id = request.POST.get('tenant_id')
        
        if tenant_id:
            try:
                tenant = Tenant.objects.get(id=tenant_id)
                # Update tenant's unit assignment
                tenant.unit = unit
                tenant.save()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': f'Tenant {tenant.first_name} {tenant.last_name} assigned to {unit.name}'})
                else:
                    return redirect('units:units_list')
            except Tenant.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Tenant not found'}, status=400)
        else:
            return JsonResponse({'success': False, 'message': 'Please select a tenant'}, status=400)
    
    # Return available tenants for modal
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        tenants = Tenant.objects.filter(unit__isnull=True)
        return JsonResponse({
            'unit_id': unit.id,
            'unit_name': unit.name,
            'tenants': list(tenants.values('id', 'first_name', 'last_name'))
        })


def detach_tenant(request, pk):
    """Detach tenant from a unit"""
    unit = get_object_or_404(Unit, pk=pk)
    
    if request.method == 'POST':
        # Find and detach the tenant assigned to this unit
        tenant = Tenant.objects.filter(unit=pk).first()
        if tenant:
            tenant.unit = None
            tenant.save()
        
        return JsonResponse({'success': True, 'message': f'Tenant detached from {unit.name}'})


@require_POST
def delete_units(request):
    """Delete selected units"""
    unit_ids = request.POST.getlist('unit_ids[]')
    
    if unit_ids:
        Unit.objects.filter(id__in=unit_ids).delete()
        return JsonResponse({'success': True, 'message': f'{len(unit_ids)} unit/s deleted successfully'})
    
    return JsonResponse({'success': False, 'message': 'No units selected'})


@require_POST
def upload_units(request):
    """Upload units from CSV or XLSX file"""
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file provided'})
    
    file = request.FILES['file']
    filename = file.name.lower()
    
    try:
        rows = []
        
        if filename.endswith('.csv'):
            # Handle CSV
            decoded_file = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            rows = list(reader)
        elif filename.endswith('.xlsx'):
            # Handle XLSX
            try:
                import openpyxl
                from openpyxl import load_workbook
                import io
                
                wb = load_workbook(io.BytesIO(file.read()))
                ws = wb.active
                
                # Get headers from first row
                headers = [cell.value for cell in ws[1]]
                
                # Convert rows to dictionaries
                for row in ws.iter_rows(min_row=2, values_only=False):
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if header:
                            row_dict[header.lower().strip()] = row[idx].value
                    if any(row_dict.values()):  # Only if row has data
                        rows.append(row_dict)
            except ImportError:
                return JsonResponse({'success': False, 'message': 'openpyxl is not installed. Please use CSV format or contact admin.'})
        else:
            return JsonResponse({'success': False, 'message': 'Invalid file format. Please use CSV or XLSX'})
        
        if not rows:
            return JsonResponse({'success': False, 'message': 'File is empty'})
        
        # Validate and create units
        created_count = 0
        errors = []
        
        for idx, row in enumerate(rows, start=2):  # start=2 because row 1 is headers
            try:
                # Normalize keys to lowercase
                row_lower = {k.lower().strip(): v for k, v in row.items()}
                
                # Check required fields
                property_name = row_lower.get('property')
                unit_name = row_lower.get('name')
                rent_amount = row_lower.get('rent_amount')
                
                if not property_name or not unit_name or rent_amount is None:
                    errors.append(f'Row {idx}: Missing required fields (property, name, rent_amount)')
                    continue
                
                # Get or create property
                try:
                    property_obj = Property.objects.get(name__iexact=property_name)
                except Property.DoesNotExist:
                    errors.append(f'Row {idx}: Property "{property_name}" not found')
                    continue
                
                # Create unit
                unit_data = {
                    'property': property_obj,
                    'name': unit_name,
                    'rent_amount': float(rent_amount) if rent_amount else 0,
                    'description': row_lower.get('description', ''),
                    'status': row_lower.get('status', 'vacant').lower(),
                }
                
                Unit.objects.create(**unit_data)
                created_count += 1
                
            except ValueError as ve:
                errors.append(f'Row {idx}: Invalid data format - {str(ve)}')
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
        
        message = f'Successfully created {created_count} units'
        if errors:
            message += f'. {len(errors)} errors: ' + '; '.join(errors[:3])
        
        return JsonResponse({'success': True, 'count': created_count, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})