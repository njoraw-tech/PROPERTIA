from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Q

from django.views.decorators.http import require_POST
from .models import Tenant
from .forms import TenantForm
from django.http import JsonResponse
from properties.models import Property
from units.models import Unit
import csv

def tenant_list(request):
    if request.method == "POST":
        form = TenantForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tenants:tenant_list')
    else:
        form = TenantForm()

    # Filtering
    tenants = Tenant.objects.all()
    property_filter = request.GET.get('property')
    status_filter = request.GET.get('status')

    if property_filter:
        # Filter by property: show tenants with units in that property, or tenants with no unit assigned
        tenants = tenants.filter(Q(unit__property_id=property_filter) | Q(unit__isnull=True))
    
    if status_filter:
        tenants = tenants.filter(status=status_filter)

    return render(request, 'tenants/tenants_view.html', {
        'tenants': tenants,
        'form': form,
        'properties': Property.objects.all(),
        'selected_property': property_filter,
        'selected_status': status_filter,
    })

@require_POST
def delete_tenants(request):
    # getlist gets all values sent under 'tenant_ids[]'
    tenant_ids = request.POST.getlist('tenant_ids[]')
    
    if not tenant_ids:
        return JsonResponse({'success': False, 'message': 'No tenants selected.'})
        
    try:
        # filter and delete
        deleted_count, _ = Tenant.objects.filter(id__in=tenant_ids).delete()
        return JsonResponse({
            'success': True, 
            'message': f'Successfully deleted {deleted_count} tenants.'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})



@require_POST
def upload_tenants(request):
    """Upload tenants from CSV or XLSX file. If 'validate_only' POST param is set, only validate and return errors without creating tenants."""
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file provided'})

    file = request.FILES['file']
    filename = file.name.lower()
    validate_only = request.POST.get('validate_only') == '1'

    try:
        rows = []

        if filename.endswith('.csv'):
            decoded_file = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            rows = list(reader)
        elif filename.endswith('.xlsx'):
            try:
                import openpyxl
                from openpyxl import load_workbook
                import io

                wb = load_workbook(io.BytesIO(file.read()))
                ws = wb.active

                headers = [cell.value for cell in ws[1]]

                for row in ws.iter_rows(min_row=2, values_only=False):
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if header:
                            row_dict[header.lower().strip()] = row[idx].value
                    if any(row_dict.values()):
                        rows.append(row_dict)
            except ImportError:
                return JsonResponse({'success': False, 'message': 'openpyxl is not installed. Please use CSV format or contact admin.'})
        else:
            return JsonResponse({'success': False, 'message': 'Invalid file format. Please use CSV or XLSX'})

        if not rows:
            return JsonResponse({'success': False, 'message': 'File is empty'})

        created_count = 0
        errors = []
        valid_rows = 0
        for idx, row in enumerate(rows, start=2):
            try:
                row_lower = {k.lower().strip(): v for k, v in row.items()}

                first_name = row_lower.get('first_name')
                last_name = row_lower.get('last_name')
                phone_number = row_lower.get('phone_number')
                property_name = row_lower.get('property')

                if not first_name or not last_name or not phone_number or not property_name:
                    errors.append(f'Row {idx}: Missing required fields (first_name, last_name, phone_number, property)')
                    continue

                # Get property object
                try:
                    property_obj = Property.objects.get(name__iexact=property_name)
                except Property.DoesNotExist:
                    errors.append(f'Row {idx}: Property "{property_name}" not found')
                    continue

                # Unit assignment is now handled by leases. Ignore unit and property fields for tenant upload.

                deposit_required = True
                if 'deposit_required' in row_lower:
                    deposit_req_val = str(row_lower.get('deposit_required', '')).lower().strip()
                    deposit_required = deposit_req_val in ['true', 'yes', '1', 't', 'y']

                deposit_amount = 0.00
                if row_lower.get('deposit_amount'):
                    try:
                        deposit_amount = float(row_lower.get('deposit_amount', 0))
                    except ValueError:
                        deposit_amount = 0.00

                if not validate_only:
                    tenant_data = {
                        'first_name': first_name,
                        'last_name': last_name,
                        'phone_number': phone_number,
                        'property': property_obj,
                        'next_of_kin_name': row_lower.get('next_of_kin_name', ''),
                        'next_of_kin_phone_number': row_lower.get('next_of_kin_phone_number', ''),
                        'description': row_lower.get('description', ''),
                        'status': row_lower.get('status', 'active').lower(),
                        'deposit_required': deposit_required,
                        'deposit_amount': deposit_amount,
                    }
                    Tenant.objects.create(**tenant_data)
                    created_count += 1
                valid_rows += 1

            except ValueError as ve:
                errors.append(f'Row {idx}: Invalid data format - {str(ve)}')
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')

        if validate_only:
            return JsonResponse({
                'success': True,
                'valid_rows': valid_rows,
                'invalid_rows': len(errors),
                'errors': errors,
                'message': f'Validation complete: {valid_rows} valid, {len(errors)} invalid.'
            })

        message = f'Successfully created {created_count} tenants'
        if errors:
            message += f'. {len(errors)} errors: ' + '; '.join(errors[:3])

        return JsonResponse({'success': True, 'count': created_count, 'message': message, 'errors': errors})

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})