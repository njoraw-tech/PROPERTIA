from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST

# Create your views here.
from django.shortcuts import render
from django.http import HttpResponse

from .models import Payment
from .forms import PaymentForm
from properties.models import Property
from tenants.models import Tenant
import csv
from datetime import datetime

def payment_list(request):
    if request.method == "POST":
        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('payments:payment_list')
    else:
        form = PaymentForm()

    # Get filters from GET parameters
    selected_property = request.GET.get('property', '')
    selected_status = request.GET.get('status', '')
    
    # Apply filters
    payments = Payment.objects.all().order_by('-date')
    
    if selected_property:
        payments = payments.filter(property_id=selected_property)
    
    if selected_status:
        payments = payments.filter(status=selected_status)
    
    properties = Property.objects.all()
    
    return render(request, 'payments/payments_view.html', {
        'payments': payments,
        'form': form,
        'properties': properties,
        'selected_property': selected_property,
        'selected_status': selected_status,
    })


@require_POST
def delete_payments(request):
    """Delete selected payments"""
    try:
        payment_ids = request.POST.getlist('payment_ids[]')
        
        if not payment_ids:
            return JsonResponse({'success': False, 'message': 'No payments selected'})
        
        deleted_count, _ = Payment.objects.filter(id__in=payment_ids).delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted {deleted_count} payment(s)'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error deleting payments: {str(e)}'
        })


@require_POST
def upload_payments(request):
    """Upload payments from CSV or XLSX file"""
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file provided'})
    
    file = request.FILES['file']
    filename = file.name.lower()
    
    try:
        rows = []
        
        if filename.endswith('.csv'):
            # Handle CSV
            decoded_file = file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file)
            rows = list(reader)
        elif filename.endswith('.xlsx'):
            # Handle XLSX
            try:
                import openpyxl
                from openpyxl import load_workbook
                import io
                
                wb = load_workbook(io.BytesIO(file.read()))
                ws = wb.active
                
                # Get headers from first row
                headers = [cell.value for cell in ws[1]]
                
                # Convert rows to dictionaries
                for row in ws.iter_rows(min_row=2, values_only=False):
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if header:
                            row_dict[header.lower().strip()] = row[idx].value
                    if any(row_dict.values()):  # Only if row has data
                        rows.append(row_dict)
            except ImportError:
                return JsonResponse({'success': False, 'message': 'openpyxl is not installed. Please use CSV format or contact admin.'})
        else:
            return JsonResponse({'success': False, 'message': 'Invalid file format. Please use CSV or XLSX'})
        
        if not rows:
            return JsonResponse({'success': False, 'message': 'File is empty'})
        
        # Validate and create payments
        created_count = 0
        errors = []
        
        for idx, row in enumerate(rows, start=2):  # start=2 because row 1 is headers
            try:
                # Normalize keys to lowercase
                row_lower = {k.lower().strip(): v for k, v in row.items()}
                
                # Check required fields
                property_name = row_lower.get('property')
                tenant_name = row_lower.get('tenant')
                amount = row_lower.get('amount')
                date_str = row_lower.get('date')
                
                if not property_name or not tenant_name or amount is None or not date_str:
                    errors.append(f'Row {idx}: Missing required fields (property, tenant, amount, date)')
                    continue
                
                # Get property
                try:
                    property_obj = Property.objects.get(name__iexact=property_name)
                except Property.DoesNotExist:
                    errors.append(f'Row {idx}: Property "{property_name}" not found')
                    continue
                
                # Get tenant
                try:
                    # Try to find tenant by first and last name or full name
                    tenant_name_str = str(tenant_name).strip()
                    if ' ' in tenant_name_str:
                        first_name, last_name = tenant_name_str.rsplit(' ', 1)
                        tenant_obj = Tenant.objects.get(first_name__iexact=first_name, last_name__iexact=last_name)
                    else:
                        tenant_obj = Tenant.objects.filter(first_name__iexact=tenant_name_str).first() or Tenant.objects.filter(last_name__iexact=tenant_name_str).first()
                        if not tenant_obj:
                            raise Tenant.DoesNotExist
                except Tenant.DoesNotExist:
                    errors.append(f'Row {idx}: Tenant "{tenant_name}" not found')
                    continue
                except:
                    errors.append(f'Row {idx}: Tenant "{tenant_name}" not found')
                    continue
                
                # Parse amount
                try:
                    amount_val = float(amount)
                except ValueError:
                    errors.append(f'Row {idx}: Invalid amount format')
                    continue
                
                # Parse date
                try:
                    if isinstance(date_str, str):
                        # Try multiple date formats
                        for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d', '%m-%d-%Y', '%d.%m.%Y']:
                            try:
                                date_obj = datetime.strptime(date_str.strip(), fmt).date()
                                break
                            except ValueError:
                                continue
                        else:
                            errors.append(f'Row {idx}: Invalid date format "{date_str}"')
                            continue
                    else:
                        date_obj = date_str
                except Exception as e:
                    errors.append(f'Row {idx}: Error parsing date - {str(e)}')
                    continue
                
                # Create payment
                payment_data = {
                    'property': property_obj,
                    'tenant': tenant_obj,
                    'amount': amount_val,
                    'date': date_obj,
                    'description': row_lower.get('description', ''),
                }
                
                Payment.objects.create(**payment_data)
                created_count += 1
                
            except ValueError as ve:
                errors.append(f'Row {idx}: Invalid data format - {str(ve)}')
            except Exception as e:
                errors.append(f'Row {idx}: {str(e)}')
        
        message = f'Successfully created {created_count} payments'
        if errors:
            message += f'. {len(errors)} errors: ' + '; '.join(errors[:3])
        
        return JsonResponse({'success': True, 'count': created_count, 'message': message})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error processing file: {str(e)}'})